"""Excel (.xlsx) parsing for bulk tender import.

Reads the first worksheet, treats row 1 as headers and yields each subsequent
row as a ``(row_number, {field: value})`` pair.

Headers are normalised and then resolved through :data:`HEADER_ALIASES`, so a
spreadsheet kept in the shape people actually maintain it — ``TENDER NO.``,
``WORK``, ``TENDER LINK`` — imports without being rewritten by hand first. The
canonical field names are always accepted; the aliases are additive.
"""

from __future__ import annotations

import io
import re
from collections.abc import Iterator
from typing import Any

from openpyxl import load_workbook

# Header text -> canonical Tender field. Keys are compared after normalisation
# (lower-cased, punctuation dropped, whitespace collapsed), so "TENDER NO." and
# "Tender No" both arrive here as "tender no".
HEADER_ALIASES: dict[str, str] = {
    # --- tender_number ---
    "tender no": "tender_number",
    "tender number": "tender_number",
    "tender id": "tender_number",
    "tender ref": "tender_number",
    "tender reference": "tender_number",
    "nit no": "tender_number",
    # --- title ---
    "work": "title",
    "work description": "title",
    "name of work": "title",
    "description of work": "title",
    "tender title": "title",
    "subject": "title",
    # --- estimated_value ---
    "value": "estimated_value",
    "tender value": "estimated_value",
    "estimated value": "estimated_value",
    "estimated cost": "estimated_value",
    "advertised value": "estimated_value",
    # --- closing_date ---
    "date": "closing_date",
    "closing date": "closing_date",
    "due date": "closing_date",
    "last date": "closing_date",
    "bid submission date": "closing_date",
    # --- source_url ---
    "tender link": "source_url",
    "link": "source_url",
    "url": "source_url",
    "source url": "source_url",
    "tender url": "source_url",
    "document link": "source_url",
    "pdf link": "source_url",
    # --- department ---
    "division": "department",
    "railway": "department",
    "zone": "department",
    "dept": "department",
    "organisation": "department",
    "organization": "department",
}

# Columns with no field of their own that are still worth keeping. They are
# folded into the description so the detail survives the import instead of
# being silently dropped, each labelled with the heading it came from.
DESCRIPTION_COLUMNS: tuple[str, ...] = (
    "location",
    "work area",
    "emd",
    "similar work",
    "status",
    "time",
    "quantity",
    "remarks",
)

_PUNCT = re.compile(r"[^a-z0-9]+")


def normalise_header(raw: object) -> str:
    """Lower-case, strip punctuation and collapse whitespace."""
    if raw is None:
        return ""
    return _PUNCT.sub(" ", str(raw).strip().lower()).strip()


def resolve_header(raw: object) -> str:
    """Map one spreadsheet heading onto a canonical field name.

    An unrecognised heading keeps its normalised form with spaces turned into
    underscores, so ``description`` and ``source_url`` written out in full still
    land on the right field without needing an alias entry.
    """
    normalised = normalise_header(raw)
    if not normalised:
        return ""
    if normalised in HEADER_ALIASES:
        return HEADER_ALIASES[normalised]
    return normalised.replace(" ", "_")


def build_description(record: dict[str, Any]) -> str | None:
    """Fold the extra columns into a single labelled description block."""
    existing = record.get("description")
    parts: list[str] = []
    if existing is not None and str(existing).strip():
        parts.append(str(existing).strip())
    for column in DESCRIPTION_COLUMNS:
        value = record.get(column.replace(" ", "_"))
        if value is None or not str(value).strip():
            continue
        parts.append(f"{column.upper()}: {str(value).strip()}")
    return "\n".join(parts) if parts else None


def parse_rows(content: bytes) -> Iterator[tuple[int, dict[str, Any]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return
        headers = [resolve_header(h) for h in header_row]
        for offset, row in enumerate(rows, start=2):
            record = {
                headers[i]: value for i, value in enumerate(row) if i < len(headers) and headers[i]
            }
            if any(v is not None and str(v).strip() for v in record.values()):
                description = build_description(record)
                if description:
                    record["description"] = description
                yield offset, record
    finally:
        workbook.close()
